"""
GeoMatrix SEO Engine
Generates GEOMATRIX keyword data and visual pitch decks from a business website + radius.
"""

import requests
from bs4 import BeautifulSoup
from anthropic import Anthropic
from geopy.geocoders import Nominatim
import json
import re
import pandas as pd
import io
import math
import time
from datetime import datetime

from PIL import Image, ImageDraw
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib.patches import Ellipse
from matplotlib.lines import Line2D

import folium
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from reportlab.pdfgen import canvas as rl_canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.colors import HexColor
from reportlab.lib.utils import ImageReader


# ─── Longtail modifier templates per broad industry ───────────────────────────
MODIFIER_TEMPLATES = {
    "default": [
        ("near me {city} {state}", "Commercial", "Money page"),
        ("cost {city} {state}", "Commercial", "Money page"),
        ("before and after {city} {state}", "Research / commercial", "Support section / FAQ"),
        ("best {city} {state}", "Commercial", "Money page"),
        ("same day {city} {state}", "Commercial", "Money page"),
        ("consultation {city} {state}", "Commercial", "Money page"),
        ("affordable {city} {state}", "Commercial", "Money page"),
        ("top rated {city} {state}", "Research / commercial", "Support section / FAQ"),
        ("licensed {city} {state}", "Research / commercial", "Support section / FAQ"),
        ("local {city} {state}", "Commercial", "Money page"),
        ("financing {city} {state}", "Commercial", "Support section / FAQ"),
        ("specialist {city} {state}", "Research / commercial", "Support section / FAQ"),
    ],
    "dental": [
        ("near me {city} {state}", "Commercial", "Money page"),
        ("cost {city} {state}", "Commercial", "Money page"),
        ("before and after {city} {state}", "Research / commercial", "Support section / FAQ"),
        ("best {city} {state}", "Commercial", "Money page"),
        ("same day {city} {state}", "Commercial", "Money page"),
        ("consultation {city} {state}", "Commercial", "Money page"),
        ("natural looking {city} {state}", "Research / commercial", "Support section / FAQ"),
        ("luxury {city} {state}", "Research / commercial", "Support section / FAQ"),
        ("for adults {city} {state}", "Research / commercial", "Support section / FAQ"),
        ("financing {city} {state}", "Commercial", "Support section / FAQ"),
        ("specialist {city} {state}", "Research / commercial", "Support section / FAQ"),
        ("affordable {city} {state}", "Commercial", "Money page"),
    ],
    "plumbing": [
        ("near me {city} {state}", "Commercial", "Money page"),
        ("cost {city} {state}", "Commercial", "Money page"),
        ("emergency {city} {state}", "Urgency", "Money page"),
        ("24/7 {city} {state}", "Urgency", "Money page"),
        ("same day {city} {state}", "Commercial", "Money page"),
        ("licensed {city} {state}", "Research / commercial", "Support section / FAQ"),
        ("best {city} {state}", "Commercial", "Money page"),
        ("affordable {city} {state}", "Commercial", "Money page"),
        ("local {city} {state}", "Commercial", "Money page"),
        ("repair {city} {state}", "Commercial", "Money page"),
        ("quotes {city} {state}", "Commercial", "Money page"),
        ("certified {city} {state}", "Research / commercial", "Support section / FAQ"),
    ],
    "hvac": [
        ("near me {city} {state}", "Commercial", "Money page"),
        ("cost {city} {state}", "Commercial", "Money page"),
        ("emergency {city} {state}", "Urgency", "Money page"),
        ("installation {city} {state}", "Commercial", "Money page"),
        ("repair {city} {state}", "Commercial", "Money page"),
        ("maintenance {city} {state}", "Commercial", "Money page"),
        ("best {city} {state}", "Commercial", "Money page"),
        ("affordable {city} {state}", "Commercial", "Money page"),
        ("licensed {city} {state}", "Research / commercial", "Support section / FAQ"),
        ("quotes {city} {state}", "Commercial", "Money page"),
        ("24/7 {city} {state}", "Urgency", "Money page"),
        ("same day {city} {state}", "Commercial", "Money page"),
    ],
    "legal": [
        ("near me {city} {state}", "Commercial", "Money page"),
        ("consultation {city} {state}", "Commercial", "Money page"),
        ("cost {city} {state}", "Commercial", "Money page"),
        ("best {city} {state}", "Commercial", "Money page"),
        ("experienced {city} {state}", "Research / commercial", "Support section / FAQ"),
        ("free consultation {city} {state}", "Commercial", "Money page"),
        ("top rated {city} {state}", "Research / commercial", "Support section / FAQ"),
        ("affordable {city} {state}", "Commercial", "Money page"),
        ("local {city} {state}", "Commercial", "Money page"),
        ("specialist {city} {state}", "Research / commercial", "Support section / FAQ"),
        ("licensed {city} {state}", "Research / commercial", "Support section / FAQ"),
        ("winning {city} {state}", "Research / commercial", "Support section / FAQ"),
    ],
}

INDUSTRY_KEYWORDS = {
    "dental": ["dental", "dentist", "orthodontic", "oral", "tooth", "teeth", "smile"],
    "plumbing": ["plumb", "pipe", "drain", "water heater", "leak", "sewer"],
    "hvac": ["hvac", "air condition", "heating", "cooling", "furnace", "heat pump"],
    "legal": ["attorney", "lawyer", "law firm", "legal", "litigation"],
}

# Industry benchmarks used for PDF revenue projections
INDUSTRY_ROI = {
    "dental":   {"searches": 50, "revenue": 1400, "conv": 0.04, "unit": "est. dental treatment value"},
    "plumbing": {"searches": 70, "revenue":  650, "conv": 0.06, "unit": "est. per service job"},
    "hvac":     {"searches": 60, "revenue":  950, "conv": 0.05, "unit": "est. per installation"},
    "legal":    {"searches": 30, "revenue": 3500, "conv": 0.03, "unit": "est. per retained client"},
    "default":  {"searches": 40, "revenue":  900, "conv": 0.04, "unit": "est. per client value"},
}


class GeomatrixEngine:

    MODEL = "claude-haiku-4-5-20251001"   # cheapest + fastest Claude model

    def __init__(self, api_key: str):
        self.client = Anthropic(api_key=api_key)
        self.geocoder = Nominatim(user_agent="geomatrix_seo_generator_v1_0", timeout=10)

    # ─── Helpers ─────────────────────────────────────────────────────────────

    def _call_claude(self, prompt: str, max_tokens: int = 4096, retries: int = 3) -> str:
        for attempt in range(retries):
            try:
                msg = self.client.messages.create(
                    model=self.MODEL,
                    max_tokens=max_tokens,
                    messages=[{"role": "user", "content": prompt}],
                )
                return msg.content[0].text
            except Exception as e:
                msg_str = str(e)
                if "overloaded" in msg_str.lower() or "529" in msg_str:
                    if attempt < retries - 1:
                        time.sleep(2 ** attempt + 2)
                        continue
                raise
        return ""

    def _extract_json(self, text: str, kind: str = "object"):
        # Try fenced code blocks first
        for pattern in [r'```json\s*(.*?)\s*```', r'```\s*(.*?)\s*```']:
            m = re.search(pattern, text, re.DOTALL)
            if m:
                try:
                    return json.loads(m.group(1))
                except Exception:
                    pass
        # Raw JSON
        bracket = r'\[.*\]' if kind == "array" else r'\{.*\}'
        m = re.search(bracket, text, re.DOTALL)
        if m:
            try:
                return json.loads(m.group())
            except Exception:
                pass
        return None

    def _detect_industry(self, industry_text: str) -> str:
        t = industry_text.lower()
        for key, kws in INDUSTRY_KEYWORDS.items():
            if any(kw in t for kw in kws):
                return key
        return "default"

    def _get_modifiers(self, industry_key: str, ai_modifiers: list = None) -> list:
        base = MODIFIER_TEMPLATES.get(industry_key, MODIFIER_TEMPLATES["default"])
        if ai_modifiers:
            extra = [(m, "Commercial", "Money page") for m in ai_modifiers if m not in [b[0] for b in base]]
            return base + extra[:4]
        return base

    # ─── Step 1: Scrape website (multi-page) ─────────────────────────────────

    # URL path fragments that suggest a services/treatments page
    SERVICE_URL_HINTS = [
        "service", "treatment", "procedure", "dental", "cosmetic", "implant",
        "veneer", "orthodont", "invisalign", "whitening", "restoration",
        "crown", "bridge", "cleaning", "emergency", "ortho", "about",
        "what-we-do", "what-we-offer", "our-services", "specialty",
        "plumb", "hvac", "roof", "legal", "law", "repair", "install",
        "contact", "location", "find-us", "directions", "office"
    ]

    def _fetch_page(self, url: str, headers: dict) -> str:
        """Fetch one page and return cleaned text."""
        try:
            resp = requests.get(url, headers=headers, timeout=12)
            resp.raise_for_status()
            soup = BeautifulSoup(resp.text, "html.parser")
            for tag in soup(["script", "style", "nav", "aside", "form"]):
                tag.decompose()
            text = soup.get_text(separator=" ", strip=True)
            return re.sub(r"\s+", " ", text)
        except Exception:
            return ""

    def _discover_service_links(self, base_url: str, html: str, limit: int = 6) -> list:
        """Find internal links that likely point to service/treatment pages."""
        from urllib.parse import urljoin, urlparse
        soup = BeautifulSoup(html, "html.parser")
        base_domain = urlparse(base_url).netloc
        seen = set()
        links = []
        for a in soup.find_all("a", href=True):
            href = a["href"].strip()
            full = urljoin(base_url, href)
            parsed = urlparse(full)
            # Must be same domain, not an anchor/file, not already seen
            if parsed.netloc != base_domain:
                continue
            path = parsed.path.lower().rstrip("/")
            if not path or path in seen or "." in path.split("/")[-1]:
                continue
            if any(hint in path for hint in self.SERVICE_URL_HINTS):
                seen.add(path)
                links.append(full)
            if len(links) >= limit:
                break
        return links

    def scrape_website(self, url: str) -> str:
        """Scrape homepage + up to 5 service/treatment sub-pages for richer content."""
        if not url.startswith(("http://", "https://")):
            url = "https://" + url
        headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/120.0.0.0"}

        # 1. Homepage
        try:
            resp = requests.get(url, headers=headers, timeout=15)
            resp.raise_for_status()
            home_html = resp.text
        except Exception as e:
            return f"[Scrape note: {e}]"

        home_text = self._fetch_page(url, headers)
        print(f"[Debug] Homepage scraped: {len(home_text)} chars")
        if len(home_text) < 300:
            print("[Debug] Warning: Scraped content is very short. Website may be JS-heavy or blocking requests.")

        # 2. Discover and scrape service sub-pages
        sub_links = self._discover_service_links(url, home_html, limit=6)
        sub_texts = []
        for link in sub_links:
            text = self._fetch_page(link, headers)
            if text:
                sub_texts.append(f"[Page: {link}]\n{text[:1500]}")
            time.sleep(0.3)   # be polite to the server

        # 3. Combine: homepage gets more room, sub-pages fill the rest
        combined = home_text[:5000]
        if sub_texts:
            combined += "\n\n--- ADDITIONAL SERVICE PAGES ---\n\n"
            combined += "\n\n".join(sub_texts)

        return combined[:12000]

    # ─── Step 2: Extract business info ───────────────────────────────────────

    def extract_business_info(self, content: str, business_name: str,
                              website: str, extra_info: str = "") -> dict:
        self.last_biz_name = business_name
        self.last_website = website
        
        extra_section = (
            f"\n\nADDITIONAL DETAILS PROVIDED BY THE USER (treat as authoritative source for location/services):\n{extra_info}"
            if extra_info and extra_info.strip() else ""
        )
        prompt = f"""Analyze this website content for "{business_name}" (website: {website}).{extra_section}

CRITICAL: You MUST find the physical address (Street, City, State). If the exact street is missing, find the City and State. Check the bottom of the text (footer) and contact mentions.

Return ONLY valid JSON (no markdown, no explanation).
{{
    "address": "full street address or 'Unknown'",
    "city": "city name or 'Unknown'",
    "state": "full state name or 'Unknown'",
    "state_abbr": "2-letter abbreviation or 'XX'",
    "zip": "zip code or empty",
    "description": "2-3 sentence business description",
    "phone": "phone number or empty",
    "industry": "e.g. Cosmetic Dentistry, Plumbing, etc.",
    "services": ["Service 1", "Service 2", "... 15-25 specific items"]
}}

Website content:
{content[:10000]}"""

        try:
            data = self._extract_json(self._call_claude(prompt), "object")
            if data:
                return data
        except Exception:
            pass
        return {
            "address": "Unknown", "city": "Unknown", "state": "Unknown",
            "state_abbr": "XX", "zip": "",
            "description": f"{business_name} provides professional services.",
            "phone": "", "industry": "Professional Services",
            "services": ["Professional Services"],
        }

    def extract_services_from_content(self, content: str, business_name: str, industry: str) -> list:
        """
        Extract every specific service/treatment from the full multi-page content.
        Merged with business info in extract_business_info to save an API call.
        This method exists for callers that need services separately.
        """
        prompt = f"""Read this website content for "{business_name}" ({industry}).

List EVERY specific service, treatment, or procedure mentioned — be maximally granular.
Each distinct service gets its own entry in Title Case.
Do NOT group (e.g. list "Porcelain Veneers" AND "No-Prep Veneers" separately, not just "Veneers").
Return 15-25 items if the content supports it.

Return ONLY a valid JSON array of strings, no markdown:
["Service 1", "Service 2", ...]

Content:
{content[:10000]}"""

        try:
            data = self._extract_json(self._call_claude(prompt), "array")
            if data and isinstance(data, list) and len(data) > 0:
                return [str(s) for s in data]
        except Exception:
            pass
        return ["Professional Services"]

    # ─── Step 3: Build Service Library ───────────────────────────────────────

    def build_service_library(self, services_list: list, industry: str,
                              business_name: str, extra_info: str = "") -> list:
        services_str = json.dumps(services_list)
        extra_note = (
            f"\nAdditional context from the business owner: {extra_info}\n"
            f"If this mentions services not already in the list, add them."
            if extra_info and extra_info.strip() else ""
        )

        prompt = f"""Build an SEO Service Library for "{business_name}" ({industry}).

Here are ALL the specific services found on their website:
{services_str}{extra_note}

Return ONLY a valid JSON array (no markdown):
[
  {{
    "Service Category": "Properly Capitalized Service Name",
    "Primary Keyword Stem": "core keyword phrase (2-4 words, what people Google)",
    "Priority": "Highest|High|Medium|Low-Med|Low",
    "Keyword Angles / Subtopics": "comma-separated content angles for this service",
    "Page Notes": "brief page strategy note",
    "CTA Angle": "action-oriented CTA e.g. Book a consult | Get a free quote | Call now"
  }}
]

Rules:
- One entry per service from the list above — do NOT merge or skip any
- Keyword stems must be the exact phrases people type in Google (e.g. "dental implants", "porcelain veneers", "teeth whitening dentist")
- Priority: Highest = highest revenue/intent (implants, veneers, etc.); Low = informational/entry (cleaning, exam)
- CTA Angle must match the service (e.g. implants → "Schedule implant consult", emergency → "Call now")
- Every entry needs a unique keyword stem
- Return one entry for EACH service in the input list"""

        try:
            data = self._extract_json(self._call_claude(prompt), "array")
            if data:
                return data
        except Exception:
            pass
        return [{"Service Category": s, "Primary Keyword Stem": s.lower(),
                 "Priority": "High", "Keyword Angles / Subtopics": s,
                 "Page Notes": "Local service page", "CTA Angle": "Get a quote"}
                for s in (services_list or ["Professional Services"])[:10]]

    # ─── Step 4: Geocode ─────────────────────────────────────────────────────

    def _geocode_via_claude(self, query: str) -> tuple:
        ctx = f"The business is '{getattr(self, 'last_biz_name', 'Unknown')}' and their website is '{getattr(self, 'last_website', 'Unknown')}'. "
        prompt = f"{ctx}Find the approximate latitude and longitude for their location based on the query '{query}'. If the query is empty or 'Unknown', use your knowledge of this business's headquarters. Return ONLY a JSON object with keys 'lat' and 'lon'. No markdown."
        try:
            data = self._extract_json(self._call_claude(prompt, max_tokens=150), "object")
            if data and "lat" in data and "lon" in data:
                return float(data["lat"]), float(data["lon"])
        except Exception:
            pass
        return None, None

    def geocode_address(self, city: str, state: str, address: str = "") -> tuple:
        city = (city or "").strip()
        state = (state or "").strip()
        address = (address or "").strip()
        
        if city == "Unknown" and state == "Unknown" and address == "Unknown":
            # Total extraction failure, try fallback immediately with context
            return self._geocode_via_claude("")

        attempts = []
        if address and address != "Unknown" and city and city != "Unknown":
            attempts.append(f"{address}, {city}, {state}")
            
        if city and state:
            attempts.append(f"{city}, {state}, USA")
            attempts.append(f"{city}, {state}")
            
        if address:
            attempts.append(address)
            
        if city:
            attempts.append(f"{city}, USA")
            attempts.append(city)
            
        for q in attempts:
            if q and "Unknown" not in q and q.strip() not in [",", ", USA", ", , USA", "USA"]:
                try:
                    loc = self.geocoder.geocode(q)
                    if loc:
                        return loc.latitude, loc.longitude
                    time.sleep(1)
                except Exception:
                    time.sleep(1)
                    
        # Fallback to AI geocoding if Nominatim fails/blocks
        fallback_query = f"{address}, {city}, {state}".replace("Unknown", "").strip(", ")
        if not fallback_query or len(fallback_query) < 3:
            fallback_query = city if city and "Unknown" not in city else address
            
        if fallback_query and "Unknown" not in fallback_query:
            lat, lon = self._geocode_via_claude(fallback_query)
            if lat and lon:
                return lat, lon
                
        return None, None

    def reverse_geocode(self, lat: float, lon: float) -> tuple:
        """Find city and state names from coordinates."""
        try:
            loc = self.geocoder.reverse((lat, lon), language="en")
            if loc and loc.raw.get("address"):
                addr = loc.raw["address"]
                city = addr.get("city") or addr.get("town") or addr.get("village") or addr.get("suburb", "Unknown")
                state = addr.get("state", "Unknown")
                return city, state
        except Exception:
            pass
        return "Unknown", "Unknown"

    # ─── Step 5: Find locations ───────────────────────────────────────────────

    def find_locations(self, lat: float, lon: float, city: str, state: str, radius_miles: int) -> list:
        phase_structure = self._get_phase_structure(radius_miles)
        if radius_miles <= 25:
            max_locs = min(40, max(20, int(radius_miles * 0.5)))
        elif radius_miles <= 50:
            max_locs = min(80, max(40, int(radius_miles * 0.2)))
        else:
            max_locs = min(250, max(80, int(radius_miles * 0.1)))

        prompt = f"""List exactly {max_locs} real cities, towns, and suburbs within {radius_miles} miles of {city}, {state}.

Return ONLY a valid JSON array (NO markdown fences). The array must contain EXACTLY {max_locs} objects:
[
  {{"name": "City Name", "distance_miles": 0.0, "market_type": "Core", "phase": "Phase 1", "lat": {lat:.4f}, "lon": {lon:.4f}}},
  ...
]

Phase and market_type rules:
{json.dumps(phase_structure, indent=2)}

Rules:
- First entry MUST be "{city}" at 0.0 miles, market_type="Core", phase="Phase 1"
- Include EXACTLY {max_locs} total entries — no more, no less
- Each entry must be a distinct, real named place (city, suburb, town, or neighborhood)
- Provide real GPS coordinates for each location
- Sort by distance_miles ascending
- Assign phase and market_type based on the distance ranges above
- DO NOT include any location more than {radius_miles} miles away"""

        try:
            data = self._extract_json(self._call_claude(prompt), "array")
            if data and isinstance(data, list):
                # Hard cap to prevent runaway responses
                return data[:max_locs]
        except Exception:
            pass
        return [{"name": city, "distance_miles": 0, "market_type": "Core",
                 "phase": "Phase 1", "lat": lat, "lon": lon}]

    def _get_phase_structure(self, radius_miles: int) -> list:
        if radius_miles <= 25:
            return [{"phase": "Phase 1", "market_type": "Core", "miles": f"0-{radius_miles}"}]
        elif radius_miles <= 50:
            return [
                {"phase": "Phase 1", "market_type": "Core", "miles": "0-25"},
                {"phase": "Phase 2", "market_type": "Expansion", "miles": f"25-{radius_miles}"},
            ]
        else:
            return [
                {"phase": "Phase 1", "market_type": "Core", "miles": "0-25"},
                {"phase": "Phase 2", "market_type": "Expansion", "miles": "25-50"},
                {"phase": "Phase 3", "market_type": "Extended", "miles": f"50-{radius_miles}"},
            ]

    @staticmethod
    def phase_configs_to_struct(phase_configs: list) -> list:
        """Convert app-level phase_configs into the display-ready phase_struct format."""
        result = []
        for pc in phase_configs:
            if pc.get("custom"):
                miles = pc["custom"][:30]
            elif pc.get("radius"):
                prev = result[-1]["miles"].split("-")[-1] if result else "0"
                miles = f"{prev}-{pc['radius']}mi"
            else:
                miles = "custom"
            result.append({"phase": pc["phase"], "market_type": pc["market_type"], "miles": miles})
        return result

    # ─── Step 5b: Multi-phase location finder ────────────────────────────────

    def find_locations_by_phases(self, lat: float, lon: float,
                                  city: str, state: str,
                                  phase_configs: list) -> list:
        """Build the full location list from per-phase configuration."""
        # Self-heal if city/state is missing (likely due to scraper blocking)
        if (not city or city == "Unknown") and lat and lon:
            city_rev, state_rev = self.reverse_geocode(lat, lon)
            city = city if city and city != "Unknown" else city_rev
            state = state if state and state != "Unknown" else state_rev

        all_locs: list = []
        seen: set = set()
        prev_radius = 0

        for i, pc in enumerate(phase_configs):
            phase      = pc["phase"]
            mtype      = pc["market_type"]
            custom     = pc.get("custom")
            radius     = pc.get("radius")

            if custom:
                locs = self._find_custom_locations(lat, lon, city, state, phase, mtype, custom)
            else:
                if i == 0:
                    n = min(80, max(30, int((radius or 0) * 0.8)))
                elif i == 1:
                    n = min(150, max(60, int((radius or 0) * 0.4)))
                else:
                    n = min(400, max(120, int((radius or 0) * 0.2)))
                
                locs = self._find_ring_locations(
                    lat, lon, city, state, phase, mtype, prev_radius, radius, n, all_locs
                )
                prev_radius = radius or prev_radius

            for loc in locs:
                if loc.get("name") and loc["name"] not in seen:
                    seen.add(loc["name"])
                    all_locs.append(loc)

        return all_locs

    def _find_ring_locations(self, lat, lon, city, state, phase, market_type,
                              prev_radius, radius, n_locs, existing_locs):
        if prev_radius == 0:
            ring_desc  = f"within {radius} miles of {city}, {state}"
            first_rule = f'First entry MUST be "{city}" at 0.0 miles'
        else:
            ring_desc  = f"between {prev_radius} and {radius} miles from {city}, {state}"
            first_rule = f"All entries should be beyond {prev_radius} miles from {city}"

        exclude    = [l["name"] for l in existing_locs]
        excl_str   = (f"\nExclude (already in earlier phase): {', '.join(exclude[:25])}"
                      if exclude else "")

        prompt = f"""List exactly {n_locs} real cities/towns/suburbs {ring_desc}.
Assign phase="{phase}", market_type="{market_type}". {first_rule}.{excl_str}

Return ONLY a valid JSON array of exactly {n_locs} objects:
[{{"name":"City","distance_miles":5.2,"market_type":"{market_type}","phase":"{phase}","lat":{lat:.4f},"lon":{lon:.4f}}}]

Rules: real named places, real GPS coords, sorted by distance ascending, within {radius} miles."""

        try:
            data = self._extract_json(self._call_claude(prompt, max_tokens=4096), "array")
            if data and isinstance(data, list):
                return data[:n_locs]
        except Exception:
            pass
        return []

    def _find_custom_locations(self, lat, lon, city, state, phase, market_type, custom_desc):
        prompt = f"""A business is based in {city}, {state}.
List up to 50 real, specific locations matching: "{custom_desc}"
Assign phase="{phase}", market_type="{market_type}".

Return ONLY a valid JSON array (up to 50 objects):
[{{"name":"Location","distance_miles":45.0,"market_type":"{market_type}","phase":"{phase}","lat":40.0,"lon":-73.0}}]

Rules: real named places only, accurate GPS coordinates, realistic distance from {city} {state}."""

        try:
            data = self._extract_json(self._call_claude(prompt, max_tokens=4096), "array")
            if data and isinstance(data, list):
                return data[:50]
        except Exception:
            pass
        return []

    # ─── Step 6: Build GeoMatrix rows ────────────────────────────────────────

    def build_geomatrix(self, service_library: list, locations: list,
                         business_name: str, state_abbr: str) -> pd.DataFrame:
        rows = []
        intent_map = {
            "Highest": "High conversion", "High": "High conversion",
            "Medium": "Support / authority", "Low-Med": "Informational", "Low": "Informational",
        }

        for loc in locations:
            city_name = loc["name"]
            dist = loc.get("distance_miles", 0)
            mtype = loc.get("market_type", "Core")
            phase = loc.get("phase", "Phase 1")

            for svc in service_library:
                stem = svc.get("Primary Keyword Stem", "").lower().strip()
                category = svc.get("Service Category", "")
                priority = svc.get("Priority", "High")
                intent = intent_map.get(priority, "Support / authority")

                primary_kw = f"{stem} {city_name} {state_abbr}"
                seo_title = f"{category} in {city_name}, {state_abbr} | {business_name}"
                if len(seo_title) > 60:
                    seo_title = f"{category} in {city_name} | {business_name}"[:60]

                # 6 longtail variants for the GeoMatrix sheet field (semicolon-separated)
                base_modifiers = [
                    f"{stem} near me {city_name}",
                    f"{stem} cost {city_name}",
                    f"{stem} best {city_name}",
                    f"{stem} same day {city_name}",
                    f"{stem} consultation {city_name}",
                    f"{stem} affordable {city_name}",
                ]
                longtail_field = "; ".join(base_modifiers)

                rows.append({
                    "Phase": phase,
                    "Location": city_name,
                    "Est. Miles From Office": round(dist, 1),
                    "Market Type": mtype,
                    "Service Category": category,
                    "Primary Keyword": primary_kw,
                    "SEO Title / Page Concept": seo_title,
                    "Longtail Variants": longtail_field,
                    "Intent": intent,
                    "Priority": priority,
                })

        return pd.DataFrame(rows)

    # ─── Step 7: Build Longtail Bank ─────────────────────────────────────────

    def build_longtail_bank(self, service_library: list, locations: list,
                             state_abbr: str, industry_key: str) -> pd.DataFrame:
        modifiers = self._get_modifiers(industry_key)
        rows = []

        for svc in service_library:
            stem = svc.get("Primary Keyword Stem", "").lower().strip()
            category = svc.get("Service Category", "")
            cta = svc.get("CTA Angle", "Get a quote")

            for loc in locations:
                city = loc["name"]
                mtype = loc.get("market_type", "Core")

                for modifier_template, intent_type, page_type in modifiers:
                    longtail_kw = f"{stem} {modifier_template}".format(
                        city=city, state=state_abbr
                    )
                    rows.append({
                        "Service Category": category,
                        "Location": city,
                        "Market Type": mtype,
                        "Base Keyword": stem,
                        "Longtail Keyword": longtail_kw,
                        "Intent Type": intent_type,
                        "Suggested Page Type": page_type,
                        "CTA Angle": cta,
                    })

        return pd.DataFrame(rows)

    # ─── Step 8: Build Page Build Plan ───────────────────────────────────────

    def build_page_plan(self, service_library: list, geomatrix_df: pd.DataFrame,
                         radius_miles: int) -> pd.DataFrame:
        phases = geomatrix_df["Phase"].unique() if "Phase" in geomatrix_df.columns else ["Phase 1"]
        high_priority = [s["Service Category"] for s in service_library
                         if s.get("Priority") in ("Highest", "High")]
        support_priority = [s["Service Category"] for s in service_library
                            if s.get("Priority") in ("Medium", "Low-Med", "Low")]

        rows = []
        for month, phase in enumerate(sorted(phases), 1):
            phase_df = geomatrix_df[geomatrix_df["Phase"] == phase] if "Phase" in geomatrix_df.columns else geomatrix_df
            locations_str = f"Top locations within {radius_miles} miles"
            page_count = len(phase_df[phase_df["Priority"].isin(["Highest", "High"])]) if "Priority" in phase_df.columns else len(phase_df)

            rows.append({
                "Month": month,
                "Phase": phase,
                "Page Group": f"Core high-value local pages ({phase})",
                "Locations": locations_str,
                "Service Focus": ", ".join(high_priority[:7]),
                "Page Count": page_count,
                "Notes": "Launch high-intent money pages first. Include internal links to service pages and new patient/client CTAs.",
            })
            if support_priority:
                rows.append({
                    "Month": month + len(phases),
                    "Phase": phase,
                    "Page Group": "Supporting service / authority pages",
                    "Locations": locations_str,
                    "Service Focus": ", ".join(support_priority[:5]),
                    "Page Count": max(len(phase_df) - page_count, 10),
                    "Notes": "Add supporting pages to build local topical authority.",
                })

        return pd.DataFrame(rows)

    # ─── Map helpers ──────────────────────────────────────────────────────────

    def _zoom_for_radius(self, radius_miles: float, lat: float, width: int = 900) -> int:
        radius_m = radius_miles * 1609.34
        for zoom in range(13, 3, -1):
            mpp = 156543.03392 * math.cos(math.radians(lat)) / (2 ** zoom)
            if radius_m / mpp <= width * 0.38:
                return zoom
        return 6

    def _latlon_to_pixel(self, lat, lon, clat, clon, zoom, W, H):
        def wx(lng): return (lng + 180) / 360 * 256 * (2 ** zoom)
        def wy(la):
            r = math.radians(la)
            return (1 - math.log(math.tan(r) + 1 / math.cos(r)) / math.pi) / 2 * 256 * (2 ** zoom)
        return int(wx(lon) - wx(clon) + W / 2), int(wy(lat) - wy(clat) + H / 2)

    # ─── Step 9a: Map image for PDF ──────────────────────────────────────────

    def generate_map_image(self, lat: float, lon: float, locations: list,
                            radius_miles: int, business_name: str) -> io.BytesIO:
        try:
            from staticmap import StaticMap, CircleMarker
            MAP_W, MAP_H = 900, 680
            zoom = self._zoom_for_radius(radius_miles, lat, MAP_W)

            m = StaticMap(MAP_W, MAP_H,
                          url_template="https://tile.openstreetmap.org/{z}/{x}/{y}.png")
            m.add_marker(CircleMarker((lon, lat), "#ff3333", 4))
            img = m.render(zoom=zoom, center=[lon, lat])
        except Exception:
            return self._matplotlib_map(lat, lon, locations, radius_miles, business_name)

        img = img.convert("RGBA")
        overlay = Image.new("RGBA", img.size, (0, 0, 0, 0))
        draw = ImageDraw.Draw(overlay)

        mpp = 156543.03392 * math.cos(math.radians(lat)) / (2 ** zoom)
        r_px = int(radius_miles * 1609.34 / mpp)
        cx, cy = MAP_W // 2, MAP_H // 2

        # Zone fills (outer → inner)
        for factor, alpha in [(1.0, 30), (0.67, 45), (0.33, 65)]:
            r = int(r_px * factor)
            draw.ellipse([cx - r, cy - r, cx + r, cy + r], fill=(0, 150, 255, alpha))
        draw.ellipse([cx - r_px, cy - r_px, cx + r_px, cy + r_px],
                     outline=(0, 180, 255, 200), width=3)

        c_map = {"Core": (0, 230, 120, 230), "Primary": (0, 230, 120, 230),
                 "Expansion": (255, 170, 0, 230), "Extended": (255, 90, 60, 230),
                 "Tertiary": (255, 90, 60, 230)}

        for loc in locations[1:]:
            if loc.get("lat") and loc.get("lon"):
                px, py = self._latlon_to_pixel(loc["lat"], loc["lon"], lat, lon, zoom, MAP_W, MAP_H)
                if 0 < px < MAP_W and 0 < py < MAP_H:
                    color = c_map.get(loc.get("market_type", "Core"), (120, 180, 255, 220))
                    draw.ellipse([px - 5, py - 5, px + 5, py + 5], fill=color)

        # Business center star
        draw.ellipse([cx - 11, cy - 11, cx + 11, cy + 11], fill=(255, 40, 40, 255),
                     outline=(255, 255, 255, 255), width=2)

        result = Image.alpha_composite(img, overlay).convert("RGB")
        buf = io.BytesIO()
        result.save(buf, format="PNG", dpi=(150, 150))
        buf.seek(0)
        return buf

    def _matplotlib_map(self, lat, lon, locations, radius_miles, business_name):
        fig = plt.figure(figsize=(9, 6.8), facecolor="#0d1b2a")
        ax = fig.add_axes([0.04, 0.04, 0.92, 0.92], facecolor="#0d1b2a")
        ax.grid(True, color="#1e3a5f", alpha=0.35, linewidth=0.5)

        lat_r = radius_miles / 69.0
        lon_r = radius_miles / (69.0 * math.cos(math.radians(lat)))

        for factor, alpha in [(1.0, 0.07), (0.67, 0.11), (0.33, 0.17)]:
            e = Ellipse((lon, lat), 2 * lon_r * factor, 2 * lat_r * factor,
                        facecolor="#00aaff", edgecolor="none", alpha=alpha, zorder=2)
            ax.add_patch(e)
        outer = Ellipse((lon, lat), 2 * lon_r, 2 * lat_r,
                        fill=False, edgecolor="#00aaff", linewidth=1.8, linestyle="--", alpha=0.6, zorder=3)
        ax.add_patch(outer)

        cm = {"Core": "#00e87a", "Expansion": "#ffaa00", "Extended": "#ff5c3a",
              "Primary": "#00e87a", "Tertiary": "#ff5c3a"}
        sm = {"Core": 70, "Expansion": 48, "Extended": 30, "Primary": 70, "Tertiary": 30}
        seen = set()
        for loc in locations[1:]:
            if loc.get("lat") and loc.get("lon") and loc["name"] not in seen:
                mt = loc.get("market_type", "Core")
                ax.scatter(loc["lon"], loc["lat"], s=sm.get(mt, 40), c=cm.get(mt, "#fff"),
                           zorder=5, alpha=0.9, edgecolors="white", linewidth=0.3)
                ax.annotate(loc["name"], (loc["lon"], loc["lat"]),
                            xytext=(3, 3), textcoords="offset points",
                            fontsize=6, color="white", alpha=0.82)
                seen.add(loc["name"])

        ax.scatter(lon, lat, s=220, c="#ff2222", marker="*", zorder=10,
                   edgecolors="white", linewidth=0.8)
        legend_els = [
            Line2D([0], [0], marker="*", color="none", markerfacecolor="#ff2222", markersize=11, label=business_name),
            Line2D([0], [0], marker="o", color="none", markerfacecolor="#00e87a", markersize=7, label="Core / Primary"),
            Line2D([0], [0], marker="o", color="none", markerfacecolor="#ffaa00", markersize=7, label="Expansion"),
            Line2D([0], [0], marker="o", color="none", markerfacecolor="#ff5c3a", markersize=7, label="Extended"),
        ]
        ax.legend(handles=legend_els, loc="lower right", facecolor="#0a1628",
                  edgecolor="#00aaff", labelcolor="white", fontsize=7.5, framealpha=0.92)
        ax.set_title(f"{radius_miles}-Mile Coverage Area", color="white", fontsize=10, pad=6)
        ax.set_xlim(lon - lon_r * 1.22, lon + lon_r * 1.22)
        ax.set_ylim(lat - lat_r * 1.22, lat + lat_r * 1.22)
        ax.set_aspect("equal")
        ax.set_xticklabels([]); ax.set_yticklabels([])
        ax.tick_params(colors="#1e3a5f")
        for sp in ax.spines.values():
            sp.set_edgecolor("#1e3a5f")

        buf = io.BytesIO()
        fig.savefig(buf, format="PNG", dpi=150, bbox_inches="tight", facecolor="#0d1b2a")
        plt.close(fig)
        buf.seek(0)
        return buf

    # ─── Step 9b: Interactive Folium map ─────────────────────────────────────

    def generate_folium_map(self, lat: float, lon: float, locations: list,
                             radius_miles: int, business_name: str) -> str:
        m = folium.Map(location=[lat, lon], zoom_start=10, tiles="CartoDB dark_matter")
        rm = radius_miles * 1609.34
        zone_styles = [
            (rm / 3, "#00aaff", 0.14, "Core Zone"),
            (2 * rm / 3, "#0088cc", 0.09, "Expansion Zone"),
            (rm, "#0066aa", 0.05, "Outer Zone"),
        ]
        for r, color, opacity, name in zone_styles:
            folium.Circle([lat, lon], radius=r, color=color, fill=True,
                          fill_opacity=opacity, popup=name, weight=1).add_to(m)
        folium.Circle([lat, lon], radius=rm, color="#00aaff",
                      fill=False, weight=2, dash_array="10").add_to(m)

        color_map = {"Core": "green", "Expansion": "orange", "Extended": "red",
                     "Primary": "green", "Tertiary": "red"}
        for loc in locations[1:]:
            if loc.get("lat") and loc.get("lon"):
                mt = loc.get("market_type", "Core")
                folium.CircleMarker(
                    [loc["lat"], loc["lon"]], radius=6,
                    color=color_map.get(mt, "blue"), fill=True, fill_opacity=0.8,
                    tooltip=loc["name"],
                    popup=f"<b>{loc['name']}</b><br>{loc.get('distance_miles', 0):.1f} mi — {mt}"
                ).add_to(m)
        folium.Marker([lat, lon], tooltip=business_name,
                      popup=f"<b>{business_name}</b><br>Business Location",
                      icon=folium.Icon(color="red", icon="star", prefix="fa")).add_to(m)
        return m._repr_html_()

    # ─── Step 10: Export Excel ────────────────────────────────────────────────

    def export_excel(self, geomatrix_df: pd.DataFrame, longtail_df: pd.DataFrame,
                     service_library: list, page_plan_df: pd.DataFrame,
                     business_name: str, website: str, business_info: dict) -> bytes:
        buf = io.BytesIO()
        wb = openpyxl.Workbook()

        h_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
        h_font = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
        fills = {
            "Core": PatternFill(start_color="c8f0da", end_color="c8f0da", fill_type="solid"),
            "Primary": PatternFill(start_color="c8f0da", end_color="c8f0da", fill_type="solid"),
            "Expansion": PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid"),
            "Extended": PatternFill(start_color="fde2da", end_color="fde2da", fill_type="solid"),
            "Tertiary": PatternFill(start_color="fde2da", end_color="fde2da", fill_type="solid"),
        }
        thin = Border(
            left=Side(style="thin", color="d0d0d0"), right=Side(style="thin", color="d0d0d0"),
            top=Side(style="thin", color="d0d0d0"), bottom=Side(style="thin", color="d0d0d0"),
        )

        def write_sheet(ws, df_or_list, col_widths=None):
            if isinstance(df_or_list, list):
                df = pd.DataFrame(df_or_list)
            else:
                df = df_or_list
            if df.empty:
                return
            for ci, col in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=ci, value=col)
                cell.fill = h_fill; cell.font = h_font; cell.border = thin
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.row_dimensions[1].height = 32
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            for ri, row in df.iterrows():
                mt = str(row.get("Market Type", ""))
                fill = fills.get(mt)
                for ci, (col, val) in enumerate(row.items(), 1):
                    cell = ws.cell(row=ri + 2, column=ci, value=str(val) if pd.notna(val) else "")
                    if fill: cell.fill = fill
                    cell.border = thin
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col_widths:
                for ci, col in enumerate(df.columns, 1):
                    w = col_widths.get(col, 18)
                    ws.column_dimensions[ws.cell(1, ci).column_letter].width = w

        # Summary sheet
        ws0 = wb.active; ws0.title = "Summary"
        summary = [
            ("Business Name", business_name),
            ("Website", website),
            ("Office Address", business_info.get("address", "")),
            ("Industry", business_info.get("industry", "")),
            ("Phone", business_info.get("phone", "")),
            ("Generated", datetime.now().strftime("%B %d, %Y")),
            ("", ""),
            ("Total GeoMatrix Rows", len(geomatrix_df)),
            ("Total Longtail Keywords", len(longtail_df)),
            ("Locations Covered", geomatrix_df["Location"].nunique() if "Location" in geomatrix_df.columns else 0),
            ("Services", geomatrix_df["Service Category"].nunique() if "Service Category" in geomatrix_df.columns else 0),
        ]
        for ri, (k, v) in enumerate(summary, 1):
            ws0.cell(ri, 1, k).font = Font(bold=True, name="Calibri")
            ws0.cell(ri, 2, str(v) if v else "").font = Font(name="Calibri")
        ws0.column_dimensions["A"].width = 28; ws0.column_dimensions["B"].width = 45

        # Service Library sheet
        ws1 = wb.create_sheet("Service Library")
        write_sheet(ws1, pd.DataFrame(service_library),
                    {"Service Category": 22, "Primary Keyword Stem": 22, "Priority": 12,
                     "Keyword Angles / Subtopics": 40, "Page Notes": 30, "CTA Angle": 28})

        # GeoMatrix sheets by phase
        phases = geomatrix_df["Phase"].unique() if "Phase" in geomatrix_df.columns else ["Phase 1"]
        phase_widths = {
            "Phase": 10, "Location": 18, "Est. Miles From Office": 10, "Market Type": 12,
            "Service Category": 22, "Primary Keyword": 32, "SEO Title / Page Concept": 38,
            "Longtail Variants": 60, "Intent": 20, "Priority": 10,
        }
        for phase in sorted(phases):
            ws = wb.create_sheet(phase)
            phase_df = geomatrix_df[geomatrix_df["Phase"] == phase] if "Phase" in geomatrix_df.columns else geomatrix_df
            write_sheet(ws, phase_df.reset_index(drop=True), phase_widths)

        # Longtail Keyword Bank
        ws_lt = wb.create_sheet("Longtail Keyword Bank")
        lt_widths = {"Service Category": 22, "Location": 18, "Market Type": 12,
                     "Base Keyword": 22, "Longtail Keyword": 42,
                     "Intent Type": 22, "Suggested Page Type": 22, "CTA Angle": 26}
        write_sheet(ws_lt, longtail_df.reset_index(drop=True), lt_widths)

        # Page Build Plan
        ws_pp = wb.create_sheet("Page Build Plan")
        write_sheet(ws_pp, page_plan_df.reset_index(drop=True),
                    {"Month": 8, "Phase": 10, "Page Group": 30, "Locations": 30,
                     "Service Focus": 50, "Page Count": 12, "Notes": 55})

        wb.save(buf); buf.seek(0)
        return buf.getvalue()

    # ─── Step 11: Export PDF ──────────────────────────────────────────────────

    def export_pdf(self, business_name: str, business_info: dict,
                   geomatrix_df: pd.DataFrame, service_library: list,
                   map_image_buf: io.BytesIO, phase_configs: list,
                   locations: list) -> bytes:
        """White, clean 1-page sales pitch deck with revenue opportunity & ROI projections."""
        buf = io.BytesIO()
        c = rl_canvas.Canvas(buf, pagesize=A4)
        W, H = A4

        # ── Design tokens (clean white theme) ──
        NAVY      = HexColor("#1a3a6b")
        BLUE      = HexColor("#2563eb")
        BLUE_LITE = HexColor("#eff6ff")
        BLUE_MID  = HexColor("#dbeafe")
        GREEN     = HexColor("#059669")
        GREEN_L   = HexColor("#d1fae5")
        ORANGE    = HexColor("#d97706")
        ORANGE_L  = HexColor("#fef3c7")
        RED_ACC   = HexColor("#dc2626")
        TEXT      = HexColor("#1e293b")
        MUTED     = HexColor("#64748b")
        BORDER    = HexColor("#e2e8f0")
        WHITE     = colors.white
        BG        = HexColor("#f8faff")   # very light blue-white background

        # ── Derived stats ──
        n_pages    = len(geomatrix_df)
        n_locs     = geomatrix_df["Location"].nunique() if "Location" in geomatrix_df.columns else len(locations)
        n_services = geomatrix_df["Service Category"].nunique() if "Service Category" in geomatrix_df.columns else len(service_library)
        n_lt       = n_pages * 12
        city_state = f"{business_info.get('city', '')}, {business_info.get('state_abbr', '')}"
        industry   = business_info.get("industry", "Services")
        industry_key  = self._detect_industry(industry)
        phase_struct  = self.phase_configs_to_struct(phase_configs)
        max_radius    = max((pc.get("radius") or 0) for pc in phase_configs) or 25
        # ROI estimates
        roi_d         = INDUSTRY_ROI.get(industry_key, INDUSTRY_ROI["default"])
        total_monthly = n_services * n_locs * roi_d["searches"]

        # ── Helpers ──
        def bg_page():
            c.setFillColor(WHITE)
            c.rect(0, 0, W, H, fill=1, stroke=0)
            # subtle top gradient strip
            c.setFillColor(BG)
            c.rect(0, H - 120, W, 120, fill=1, stroke=0)

        def rule(y, col=BORDER, lw=0.8):
            c.setStrokeColor(col); c.setLineWidth(lw)
            c.line(28, y, W - 28, y)

        def card_rect(x, y, w, h, fill=BLUE_LITE, border=BORDER, radius=5):
            c.setFillColor(fill)
            c.roundRect(x, y, w, h, radius, fill=1, stroke=0)
            if border:
                c.setStrokeColor(border); c.setLineWidth(0.6)
                c.roundRect(x, y, w, h, radius, fill=0, stroke=1)

        def label(text, x, y, size=8, col=MUTED, bold=False, align="left"):
            c.setFillColor(col)
            c.setFont("Helvetica-Bold" if bold else "Helvetica", size)
            if align == "center": c.drawCentredString(x, y, text)
            elif align == "right": c.drawRightString(x, y, text)
            else: c.drawString(x, y, text)

        def stat_block(x, y, w, h, number, caption, accent=BLUE):
            card_rect(x, y, w, h, WHITE, BORDER)
            # Accent left bar
            c.setFillColor(accent); c.rect(x, y, 4, h, fill=1, stroke=0)
            c.setFillColor(accent); c.setFont("Helvetica-Bold", 20)
            c.drawCentredString(x + w / 2 + 2, y + h - 28, str(number))
            c.setFillColor(MUTED); c.setFont("Helvetica", 7)
            c.drawCentredString(x + w / 2 + 2, y + 8, caption.upper())

        def bullet_check(x, y, text, col=GREEN, text_col=TEXT, size=8):
            c.setFillColor(col); c.setFont("Helvetica-Bold", size)
            c.drawString(x, y, "✓")
            c.setFillColor(text_col); c.setFont("Helvetica", size)
            c.drawString(x + 12, y, text)

        def stat_row(x, y, icon, headline, sub, accent=BLUE):
            c.setFillColor(accent); c.setFont("Helvetica-Bold", 10)
            c.drawString(x, y, icon)
            c.setFillColor(TEXT); c.setFont("Helvetica-Bold", 8.5)
            c.drawString(x + 16, y, headline)
            c.setFillColor(MUTED); c.setFont("Helvetica", 7.5)
            c.drawString(x + 16, y - 11, sub)

        # ════════════════ PAGE 1 ════════════════
        bg_page()

        # ── Top navy bar + header ──────────────────────────────────────────
        c.setFillColor(NAVY); c.rect(0, H - 6, W, 6, fill=1, stroke=0)

        HDR_Y = H - 6 - 68
        # Left: logo circle + business name
        cx_icon = 50; cy_icon = H - 6 - 34
        c.setFillColor(BLUE); c.circle(cx_icon, cy_icon, 18, fill=1, stroke=0)
        c.setFillColor(WHITE); c.setFont("Helvetica-Bold", 12)
        initials = "".join(w[0].upper() for w in business_name.split()[:2])
        c.drawCentredString(cx_icon, cy_icon - 4, initials)

        biz_display = business_name if len(business_name) <= 40 else business_name[:38] + "…"
        label(biz_display, 78, H - 6 - 26, size=18, col=NAVY, bold=True)
        label(f"Local SEO Growth Strategy  ·  {city_state}  ·  {industry}", 78, H - 6 - 44, size=8.5, col=MUTED)

        # Right: date + GeoMatrix badge
        label(datetime.now().strftime("%B %d, %Y"), W - 28, H - 6 - 24, size=8, col=MUTED, align="right")
        label("Powered by GeoMatrix™", W - 28, H - 6 - 38, size=7.5, col=BLUE, align="right")

        # Thin divider + tagline
        rule(H - 6 - 70)
        label("  Capturing Your Local Market — Systematically.", 28, H - 6 - 83, size=10.5, col=NAVY, bold=True)

        # ── Main content: Map (left) + Pitch panel (right) ────────────────
        MAP_X  = 28;    MAP_W  = 296
        INFO_X = 336;   INFO_W = W - INFO_X - 28
        BODY_TOP = H - 6 - 92
        MAP_H  = 252

        # Map card
        card_rect(MAP_X - 2, BODY_TOP - MAP_H - 2, MAP_W + 4, MAP_H + 4,
                  fill=BLUE_LITE, border=BORDER)
        if map_image_buf:
            map_image_buf.seek(0)
            ir = ImageReader(map_image_buf)
            c.drawImage(ir, MAP_X, BODY_TOP - MAP_H,
                        width=MAP_W, height=MAP_H, preserveAspectRatio=True, mask="auto")

        # Map caption
        phases_label = " + ".join(pc["phase"] for pc in phase_configs)
        label(f"Service Area Coverage — {phases_label}", MAP_X + MAP_W / 2,
              BODY_TOP - MAP_H - 14, size=7.5, col=MUTED, align="center")

        # ── Right panel: Market Opportunity + Revenue Projection + Services ─
        ry = BODY_TOP

        # ── Box 1: Market Opportunity ──────────────────────────────────────
        BOX_H1 = 110
        card_rect(INFO_X, ry - BOX_H1, INFO_W, BOX_H1, BLUE_LITE, BORDER)
        label("MARKET OPPORTUNITY", INFO_X + 10, ry - 14, size=7.5, col=BLUE, bold=True)

        # Big search-volume number
        vol_str = f"~{total_monthly:,}"
        c.setFillColor(NAVY); c.setFont("Helvetica-Bold", 22)
        c.drawString(INFO_X + 10, ry - 40, vol_str)
        c.setFillColor(MUTED); c.setFont("Helvetica", 7)
        c.drawString(INFO_X + 10, ry - 52, "estimated monthly searches in your service area")

        # Horizontal rule + compact stats
        c.setStrokeColor(BORDER); c.setLineWidth(0.5)
        c.line(INFO_X + 6, ry - 58, INFO_X + INFO_W - 6, ry - 58)
        compact_stats = [
            "46% of all Google searches have local intent",
            "Top 3 positions capture 68% of available clicks",
            "14.6% local close rate vs. 1.7% from paid ads",
        ]
        sy = ry - 70
        for stat in compact_stats:
            c.setFillColor(BLUE); c.setFont("Helvetica-Bold", 7)
            c.drawString(INFO_X + 10, sy, "•")
            c.setFillColor(TEXT); c.setFont("Helvetica", 7)
            c.drawString(INFO_X + 20, sy, stat)
            sy -= 12

        ry -= BOX_H1 + 6

        # ── Box 2: Revenue Projection ──────────────────────────────────────
        BOX_H2 = 98
        card_rect(INFO_X, ry - BOX_H2, INFO_W, BOX_H2, fill=WHITE, border=BORDER)
        label("REVENUE PROJECTION", INFO_X + 10, ry - 13, size=7.5, col=BLUE, bold=True)

        scenarios = [
            ("Conservative", 0.03, MUTED),
            ("Moderate",     0.07, NAVY),
            ("Aggressive",   0.12, GREEN),
        ]
        col_w3 = INFO_W / 3
        for ci, (sc_lbl, cap, sc_col) in enumerate(scenarios):
            cx = INFO_X + col_w3 * ci + col_w3 / 2
            visitors   = int(total_monthly * cap)
            leads      = max(1, int(visitors * roi_d["conv"]))
            annual_rev = leads * 12 * roi_d["revenue"]
            rev_str    = (f"${annual_rev/1_000_000:.1f}M/yr"
                          if annual_rev >= 1_000_000 else f"${annual_rev/1000:.0f}K/yr")

            c.setFillColor(sc_col); c.setFont("Helvetica-Bold", 7)
            c.drawCentredString(cx, ry - 26, sc_lbl)

            c.setFillColor(sc_col); c.setFont("Helvetica-Bold", 9)
            c.drawCentredString(cx, ry - 41, f"{int(cap*100)}% capture")

            c.setFillColor(TEXT); c.setFont("Helvetica", 7.5)
            c.drawCentredString(cx, ry - 56, f"{leads:,} leads/mo")

            c.setFillColor(sc_col); c.setFont("Helvetica-Bold", 9.5 if ci == 1 else 8.5)
            c.drawCentredString(cx, ry - 72, rev_str)

        # Vertical col dividers
        c.setStrokeColor(BORDER); c.setLineWidth(0.5)
        for di in [1, 2]:
            dx = INFO_X + col_w3 * di
            c.line(dx, ry - BOX_H2 + 6, dx, ry - 18)

        # Footnote
        c.setFillColor(MUTED); c.setFont("Helvetica", 6)
        c.drawString(INFO_X + 6, ry - 87, f"Model: {roi_d['unit']} · industry avg conv. rate")

        ry -= BOX_H2 + 6

        # Services covered (compact 2-column list, 8 items max)
        svc_names = [s["Service Category"] for s in service_library[:8]]
        # 8 items in 2 cols = 4 rows; rows drawn at ry-26, ry-48, ry-70, ry-92
        svc_box_h = 14 + math.ceil(len(svc_names) / 2) * 22 + 8
        card_rect(INFO_X, ry - svc_box_h, INFO_W, svc_box_h, fill=WHITE, border=BORDER)
        label("SERVICES TARGETED", INFO_X + 10, ry - 13, size=7.5, col=BLUE, bold=True)
        sy2 = ry - 26
        for i, svc in enumerate(svc_names):
            col_x = INFO_X + 10 if i % 2 == 0 else INFO_X + INFO_W / 2
            if i % 2 == 0 and i > 0:
                sy2 -= 11
            c.setFillColor(BLUE); c.setFont("Helvetica-Bold", 7)
            c.drawString(col_x, sy2, "•")
            c.setFillColor(TEXT); c.setFont("Helvetica", 7)
            c.drawString(col_x + 8, sy2, svc[:24])
            if i % 2 == 1:
                sy2 -= 11

        # ── Stat cards row ────────────────────────────────────────────────
        STATS_Y   = BODY_TOP - MAP_H - 30
        card_w    = (MAP_W + 4 - 3 * 6) / 4
        stats_row = [
            (f"{n_pages:,}", "SEO Pages", BLUE),
            (f"{n_lt:,}", "Keywords", GREEN),
            (str(n_locs), "Locations", ORANGE),
            (f"{max_radius}mi", "Max Radius", NAVY),
        ]
        for i, (val, cap, acc) in enumerate(stats_row):
            stat_block(MAP_X + i * (card_w + 6), STATS_Y - 52, card_w, 52, val, cap, acc)

        # ── Phase rollout ─────────────────────────────────────────────────
        PH_Y = STATS_Y - 52 - 18
        rule(PH_Y + 4, NAVY, 0.5)
        label("THE PHASED ROLLOUT PLAN", 28, PH_Y - 8, size=8, col=NAVY, bold=True)

        # Phase descriptors — only show the phases that exist for this radius
        _all_ph_descs = [
            ("PHASE 1", "Core Market",       "Highest-intent pages\nfor your backyard.\nDrive early wins fast.", GREEN),
            ("PHASE 2", "Market Expansion",  "Grow reach to surrounding\ncommunities and\nsecondary markets.",   BLUE),
            ("PHASE 3", "Authority & Scale", "Full regional coverage.\nDominate every search\nin your service area.", ORANGE),
        ]
        n_phases  = len(phase_struct)
        ph_descs  = _all_ph_descs[:n_phases]
        ph_total_w = MAP_W + 4
        ph_gap     = 8
        ph_box_w   = (ph_total_w - ph_gap * (n_phases - 1)) / n_phases
        ph_box_h   = 90
        PH_BOX_Y   = PH_Y - 24 - ph_box_h

        for i, (ph_name, ph_sub, ph_body, ph_col) in enumerate(ph_descs):
            bx = MAP_X + i * (ph_box_w + 8)
            by_box = PH_BOX_Y

            # Outer card
            card_rect(bx, by_box, ph_box_w, ph_box_h, fill=WHITE, border=BORDER)
            # Colored top header (26px — wide enough for large phase number)
            c.setFillColor(ph_col)
            c.roundRect(bx, by_box + ph_box_h - 26, ph_box_w, 26, 5, fill=1, stroke=0)
            c.rect(bx, by_box + ph_box_h - 26, ph_box_w, 12, fill=1, stroke=0)

            # Large phase number on left of header
            c.setFillColor(WHITE); c.setFont("Helvetica-Bold", 17)
            c.drawString(bx + 7, by_box + ph_box_h - 18, str(i + 1))
            # Phase name to the right of the number
            c.setFillColor(WHITE); c.setFont("Helvetica-Bold", 8)
            c.drawString(bx + 26, by_box + ph_box_h - 13, ph_name)

            # Subtitle
            c.setFillColor(ph_col); c.setFont("Helvetica-Bold", 8)
            c.drawCentredString(bx + ph_box_w / 2, by_box + ph_box_h - 38, ph_sub)

            # Body lines — no circle at bottom, text has clear space
            for j, line in enumerate(ph_body.split("\n")):
                c.setFillColor(TEXT); c.setFont("Helvetica", 7.2)
                c.drawCentredString(bx + ph_box_w / 2, by_box + ph_box_h - 52 - j * 11, line.strip())

        # Phase on right side (months indicator)
        ph_right_y = PH_Y - 26
        for i, ph in enumerate(phase_struct):
            month_start = i * 2 + 1
            month_end   = month_start + 1
            label(f"{ph['phase']} · {ph['miles']} mi · Month {month_start}–{month_end}",
                  INFO_X, ph_right_y - i * 16, size=7.5, col=TEXT if i == 0 else MUTED)

        # ── Value proposition strip ───────────────────────────────────────
        VP_Y = PH_BOX_Y - 14
        rule(VP_Y + 4, BORDER)

        vp_items = [
            (BLUE,   "Search-Ready",  "Pages built around exactly\nwhat your customers type"),
            (GREEN,  "Phase-Tracked", "Measurable ROI at every\nstage of the rollout"),
            (ORANGE, "Full Coverage", "No competitor owns your\nmarket — you will"),
        ]
        vp_box_h = 60
        vp_box_w = (W - 56 - 2 * 10) / 3
        VP_BOX_Y = VP_Y - 12 - vp_box_h

        for i, (col, title, body) in enumerate(vp_items):
            bx = 28 + i * (vp_box_w + 10)
            card_rect(bx, VP_BOX_Y, vp_box_w, vp_box_h, fill=BLUE_LITE if col == BLUE else GREEN_L if col == GREEN else ORANGE_L, border=col)
            c.setFillColor(col); c.setFont("Helvetica-Bold", 9)
            c.drawString(bx + 10, VP_BOX_Y + vp_box_h - 18, title)
            for j, line in enumerate(body.split("\n")):
                c.setFillColor(TEXT); c.setFont("Helvetica", 7.5)
                c.drawString(bx + 10, VP_BOX_Y + vp_box_h - 34 - j * 12, line.strip())

        # ── CTA footer ───────────────────────────────────────────────────
        CTA_Y = VP_BOX_Y - 12
        c.setFillColor(NAVY)
        c.roundRect(28, CTA_Y - 36, W - 56, 36, 5, fill=1, stroke=0)

        cta_text = "Ready to dominate local search? Let's start with Phase 1."
        c.setFillColor(WHITE); c.setFont("Helvetica-Bold", 11)
        c.drawString(42, CTA_Y - 22, cta_text)

        phone = business_info.get("phone", "")
        if phone:
            c.setFillColor(HexColor("#93c5fd")); c.setFont("Helvetica", 9)
            c.drawRightString(W - 42, CTA_Y - 22, phone)

        # Page footer
        c.setFillColor(MUTED); c.setFont("Helvetica", 7)
        c.drawString(28, 14, f"GeoMatrix Local SEO Strategy  ·  {business_name}  ·  Confidential")
        c.drawRightString(W - 28, 14, datetime.now().strftime("%B %Y"))

        c.save(); buf.seek(0)
        return buf.getvalue()
